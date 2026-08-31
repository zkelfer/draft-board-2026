"""Extract raw stat projections from data_private/DraftSheets_2026.xlsx -> proj.json.

Position tabs hold per-player stat triplets (avg row, then 'high'/'low' variant rows).
ECR tab supplies projected missed games. Keys match build_data.py's norm() so the
template can join proj.json onto data.json by player id.
Run before build.py. If the xlsx is absent, writes proj.json as null.
"""
import json, re, pathlib
here = pathlib.Path(__file__).parent
XLSX = here.parent/"data_private"/"DraftSheets_2026.xlsx"

def norm_key(name):
    n = name.replace("’","'").replace("‘","'")
    k = n.lower().replace(".","").replace("'","")
    k = re.sub(r"\b(jr|sr|iii|ii)\b","",k)
    k = re.sub(r"\s+"," ",k).strip()
    k = {"kenny gainwell":"kenneth gainwell","chig okonkwo":"chigoziem okonkwo"}.get(k,k)
    return k

if not XLSX.exists():
    (here/"proj.json").write_text("null")
    print("no xlsx at", XLSX, "- wrote null proj.json")
    raise SystemExit

import openpyxl
wb = openpyxl.load_workbook(XLSX, data_only=True)

# stat column layout per tab (1-indexed, after Player/Team)
LAYOUT = {
  "QB": ["paATT","paCMP","paYDS","paTD","paINT","ruATT","ruYDS","ruTD","FL"],
  "RB": ["ruATT","ruYDS","ruTD","rec","reYDS","reTD","FL"],
  "WR": ["rec","reYDS","reTD","ruATT","ruYDS","ruTD","FL"],
  "TE": ["rec","reYDS","reTD","FL"],
}

missed, bye = {}, {}
ws = wb["ECR"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[2]:  # PLAYER NAME col C: bye week
        try: bye[norm_key(str(r[2]))] = int(r[5])
        except (TypeError, ValueError): pass
    if r[11]: # PLAYER col L: proj missed games
        try: missed[norm_key(str(r[11]))] = round(float(r[12]), 2)
        except (TypeError, ValueError): pass

players = {}
for pos, cols in LAYOUT.items():
    ws = wb[pos]
    cur = None
    for r in ws.iter_rows(min_row=2, values_only=True):
        name, second = r[0], r[1]
        stats = {c: round(float(v),1) for c,v in zip(cols, r[2:2+len(cols)]) if v is not None}
        if name and str(name).strip() and str(name).strip() != "\xa0":
            key = norm_key(str(name).strip())
            cur = {"pos":pos, "team":second, "a":stats}
            players[key] = cur
        elif cur is not None and second in ("high","low"):
            cur["h" if second=="high" else "l"] = stats
        else:
            cur = None

from scratches import SCRATCHED
for k in SCRATCHED: players.pop(k, None)

for k,p in players.items():
    p["m"] = missed.get(k, 0)
    if k in bye: p["bye"] = bye[k]

out = {"players": players}
json.dump(out, open(here/"proj.json","w"), separators=(",",":"))
have_hl = sum(1 for p in players.values() if "h" in p and "l" in p)
print(f"proj.json: {len(players)} players "
      f"({have_hl} with high/low bands, {sum(1 for p in players.values() if p['m'])} with missed-game proj)")
